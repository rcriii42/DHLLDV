"""test_excel_import - Tests of load_pump_excel in the Viewer

Added by R. Ramsdell 2024-06-20
"""
import unittest

import openpyxl

from DHLLDV_viewer import load_pump_excel


class MyTestCase(unittest.TestCase):
    def test_load_objects(self):
        """Test loading the pipeline and pump objects correctly"""
        fname = "tests/Example_excel_input.xlsx"
        wb = openpyxl.load_workbook(filename=fname, data_only=True)
        pipeline = load_pump_excel.load_pipeline_from_workbook(wb)
        self.assertEqual('UWP 34x34x60 @ 300RPM', pipeline.pumps[0].name)
        self.assertEqual('0.864x0.864x2.134m Pump at 315 RPM', pipeline.pumps[1].name)
        self.assertEqual('Entrance', pipeline.pipesections[0].name)
        self.assertEqual('UWP 34x34x60 @ 300RPM', pipeline.pipesections[2].name)
        self.assertEqual('Shore Line', pipeline.pipesections[-1].name)

    def test_load_calculations(self):
        """Test that the loaded pipeline calculations are correct"""
        fname = "tests/Example_excel_input.xlsx"
        wb = openpyxl.load_workbook(filename=fname, data_only=True)
        pipeline = load_pump_excel.load_pipeline_from_workbook(wb)

        flow_list = [pipeline.pipesections[-1].flow(v) for v in pipeline.slurry.vls_list]
        self.assertAlmostEqual(3.4017658516096114, pipeline.find_operating_point(flow_list))

        flow = 4.0
        self.assertTupleEqual((155.83833315609024, 147.87084371417842, 101.75436068325409, 95.93219789540075),
                              pipeline.calc_system_head(flow))
        self.assertTupleEqual((4.0, 74.32723942350866, 3728.499999999634, 4.811435751810117),
                              pipeline.pumps[1].point(flow))

    def test_load_missing_driver_table(self):
        """Test catching missing table in loaded Excel file power tab"""
        fname = "tests/Example_excel_input_NoDriverTable.xlsx"
        wb = openpyxl.load_workbook(filename=fname, data_only=True)
        self.assertRaises(load_pump_excel.InvalidExcelError, load_pump_excel.load_pipeline_from_workbook, wb)

    def test_load_missing_pump_column(self):
        """Test catching missing column in loaded Excel file pump table"""
        fname = "tests/Example_excel_input_MissingColumn.xlsx"
        wb = openpyxl.load_workbook(filename=fname, data_only=True)
        self.assertRaises(load_pump_excel.InvalidExcelError, load_pump_excel.load_pipeline_from_workbook, wb)

    def test_load_extra_pipeline_column(self):
        """Test catching duplicate column in loaded Excel file pipeline table"""
        fname = "tests/Example_excel_input_ExtraColumn.xlsx"
        wb = openpyxl.load_workbook(filename=fname, data_only=True)
        self.assertRaises(load_pump_excel.InvalidExcelError, load_pump_excel.load_pipeline_from_workbook, wb)


if __name__ == '__main__':
    unittest.main()
