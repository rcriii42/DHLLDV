"""Store pipelines to Excel

Create an excel file with the format defines in load_pump_excel

Added by R. Ramsdell 2023-07-02"""

from datetime import datetime
import os.path
import string

import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname, absolute_coordinate, get_column_letter

from DHLLDV.DriverObj import Driver
from DHLLDV.PumpObj import Pump
from DHLLDV.PipeObj import Pipe, Pipeline
from DHLLDV.SlurryObj import Slurry
try:
    from load_pump_excel import excel_requireds
    from load_pump_excel import __doc__ as xl_doc
except ImportError:
    from .load_pump_excel import excel_requireds
    from .load_pump_excel import __doc__ as xl_doc


# These are the characters allowed in filenames when saving to excel
replace_filename_chars = [' ', ]  # These will be replaced with '_', applied before deleting invalid characters
valid_filename_chars = f'-_{string.ascii_letters}{string.digits}'


def remove_disallowed_filename_chars(filename_candidate: str, extension: str or None = None) -> str:
    """Turn a string into an acceptable filename using a whitelist approach

    The characters in replace_filename_characters are replaced with '_', other unwanted characters are removed

    filename_candidate is the string to be transformed
    extension is the filename extension to add

    Returns the cleaned filename
    """
    if extension is None:
        extension = ''
    for c in replace_filename_chars:
        cleaned_filename = filename_candidate.replace(c, '_')
    cleaned_filename = ''.join(c for c in cleaned_filename if c in valid_filename_chars)
    cleaned_filename += extension
    return cleaned_filename


def create_and_fill_named_range(wb: openpyxl.Workbook, sheet_name: str, range_name: str, range_addr: str,
                                value: str or float) -> None:
    """Create a worksheet-scope named range and fill with the given value

    wb: The workbook
    sheet_name: The sheet with the range
    range_name: The name of the defined name, must meet excel rules
    range_addr: The cell address in excel "A1" format
    value: The value to place in the cell
    """
    ws = wb[sheet_name]
    ws[range_addr].value = value
    ref = f"{quote_sheetname(ws.title)}!{absolute_coordinate(range_addr)}"
    defn = DefinedName(range_name, attr_text=ref)
    ws.defined_names.add(defn)


def write_slurry_to_excel(wb: openpyxl.workbook, slurry: Slurry, reqs: dict) -> None:
    """Write a slurry to the workbook

    wb: The Workbook
    slurry: The Slurry to write
    reqs: The 'slurry' dict from excel_requireds
    """
    sheet_name = 'slurry'
    ws = wb.create_sheet(sheet_name)
    current_row = 1
    units_map = {'name': '',
                 'pipe_dia': 'm',
                 'd_15': 'mm',
                 'd_50': 'mm',
                 'd_85': 'mm',
                 'fluid': 'fresh/salt',
                 'Cv': '-',
                 'rhos': "ton/m3",
                 'rhoi': "ton/m3"}
    for range_name, val_type in reqs.items():
        if range_name == 'required':
            continue
        elif 'd_' in range_name:
            value = slurry.get_dx(float(range_name[2:])/100)*1000
        elif range_name == 'pipe_dia':
            value = slurry.Dp
        else:
            try:
                value = slurry.__dict__[range_name]
            except KeyError:
                value = slurry.__dict__[f'_{range_name}']  # Several parameters have shadow "_xxx" variables
        ws[f'A{current_row}'].value = range_name
        create_and_fill_named_range(wb, sheet_name, range_name, f'B{current_row}', value)
        ws[f'C{current_row}'].value = units_map[range_name]
        current_row += 1


def write_driver_to_excel(wb: openpyxl.Workbook, this_driver: Driver, driver_name: str, reqs: dict):
    ws = wb.create_sheet(driver_name)
    current_row = 1
    units_map = {'name': '',
                 }
    for range_name, val_type in reqs.items():
        if range_name in ('required', 'power_curve'):
            continue
        else:
            try:
                value = this_driver.__dict__[range_name]
            except KeyError:
                value = this_driver.__dict__[f'_{range_name}']  # Several parameters have shadow "_xxx" variables
        ws[f'A{current_row}'].value = range_name
        create_and_fill_named_range(wb, driver_name, range_name, f'B{current_row}', value)
        ws[f'C{current_row}'].value = units_map[range_name]
        current_row += 2

    # The speed/power curve
    current_col = 1
    first_row = current_row
    for header in ('Speed Hz', 'Power kW'):
        ws.cell(row=current_row, column=current_col).value = header
        current_col += 1
    current_row += 1
    for speed, power in this_driver.design_power_curve.items():
        current_col = 1
        for value in (speed, power):
            ws.cell(row=current_row, column=current_col).value = value
            current_col += 1
        current_row += 1
    range_addr = f'A{first_row}:{get_column_letter(current_col - 1)}{current_row - 1}'
    ref = f"{quote_sheetname(ws.title)}!{absolute_coordinate(range_addr)}"
    defn = DefinedName('power_curve', attr_text=ref)
    ws.defined_names.add(defn)


def write_pump_to_excel(wb: openpyxl.Workbook, this_pump: Pump, pump_name: str, pump_reqs: dict, driver_reqs: dict):
    ws = wb.create_sheet(pump_name)
    current_row = 1
    units_map = {'name': '',
                 'design_impeller': 'm',
                 'suction_dia': 'm',
                 'disch_dia': 'm',
                 'design_speed': 'Hz',
                 'limited': 'torque/power/curve',
                 'gear_ratio': '-',
                 'avail_power': 'kW'}
    for range_name, val_type in pump_reqs.items():
        if val_type in [str, int, float]:
            try:
                value = this_pump.__dict__[range_name]
            except KeyError:
                value = this_pump.__dict__[f'_{range_name}']  # Several parameters have shadow "_xxx" variables
            ws[f'A{current_row}'].value = range_name
            create_and_fill_named_range(wb, pump_name, range_name, f'B{current_row}', value)
            ws[f'C{current_row}'].value = units_map[range_name]
            current_row += 1

    # The flow/head/power curves
    current_col = 1
    first_row = current_row
    for header in ('Flow m3/sec', 'Head m', 'Power kW'):
        ws.cell(row=current_row, column=current_col).value = header
        current_col += 1
    current_row += 1
    for point in zip(this_pump.design_QH_curve.keys(),
                     this_pump.design_QH_curve.values(),
                     this_pump.design_QP_curve.values(),
                     this_pump.design_QP_curve.keys()):
        current_col = 1
        # TODO: Check here for different QH/QP flow lists
        for value in point[:-1]:
            ws.cell(row=current_row, column=current_col).value = value
            current_col += 1
        current_row += 1
    range_addr = f'A{first_row}:{get_column_letter(current_col - 1)}{current_row - 1}'
    ref = f"{quote_sheetname(ws.title)}!{absolute_coordinate(range_addr)}"
    defn = DefinedName('pump_curve', attr_text=ref)
    ws.defined_names.add(defn)

    # The driver
    if this_pump.limited == 'curve':
        driver_name = pump_name.replace('Pump', 'Driver')
        write_driver_to_excel(wb, this_pump.driver, driver_name, driver_reqs)


def write_pipesections_to_excel(wb: openpyxl.workbook, pipesections: list[Pipe, Pump],
                                current_row: int, requireds: dict or None = None) -> int:
    """Write the pipe section table to the workbook

    wb: The Workbook
    pipesections: The list of Pipe or Pump objects
    current_row: The row to start writing
    requireds: the Dict of required names, uses excel_requireds if None

    returns the new current_row
    """
    if requireds is None:
        requireds = excel_requireds
    sheet_name = 'pipeline'
    ws = wb[sheet_name]
    current_col = 1
    first_row = current_row
    for header in ('Pipe Name', 'Diameter (m)', 'Length (m)', 'Total K (-)', 'Elev Change (m)', 'Final Elev (m)'):
        ws.cell(row=current_row, column=current_col).value = header
        current_col += 1
    current_row += 1
    current_elev = 0
    pump_num = 1
    for p in pipesections:
        if isinstance(p, Pump):
            pump_name = f'Number{pump_num}Pump'
            write_pump_to_excel(wb, p, pump_name, requireds['pump'], requireds['driver'])
            current_col = 1
            for value in (pump_name, '', '', '', '', current_elev):
                ws.cell(row=current_row, column=current_col).value = value
                current_col += 1
            pump_num += 1
        elif isinstance(p, Pipe):
            current_elev += p.elev_change
            current_col = 1
            for value in (p.name, p.diameter, p.length, p.total_K, p.elev_change, current_elev):
                ws.cell(row=current_row, column=current_col).value = value
                current_col += 1
        else:
            raise TypeError(f'Object of type {type(p)} not supported when storing to Excel')
        current_row += 1
    range_addr = f'A{first_row}:{get_column_letter(current_col - 1)}{current_row - 1}'
    ref = f"{quote_sheetname(ws.title)}!{absolute_coordinate(range_addr)}"
    defn = DefinedName('pipe_table', attr_text=ref)
    ws.defined_names.add(defn)

    return current_row


def store_to_excel(pipeline: Pipeline, fname: str or None = None, requireds: dict or None = None,
                   path: str or None = None) -> str:
    """Store the pipeline to a new Excel file

    pipeline: The pipeline to store
    fname: The name of the file to store. If None, the filename will be a version of the pipeline name with a timestamp
    requireds: A dict with the layout of the file. If None, use excel_requireds defined above
    path: The folder path (relative to .) for saving

    returns the new filename
    """
    if requireds is None:
        requireds = excel_requireds
    if path is None:
        path = os.path.join(os.path.dirname(__file__), 'static', 'pipelines')
    if fname is None:
        basename = f'{pipeline.name}_{datetime.now().strftime("%Y-%m-%d_%H%M%S")}'
        fname = os.path.join(path, remove_disallowed_filename_chars(basename, ".xlsx"))
    else:
        if '.xlsx' in fname and fname[-5:] == '.xlsx':
            fname = os.path.join(path, remove_disallowed_filename_chars(fname[:-5], '.xlsx'))
        else:
            fname = os.path.join(path, remove_disallowed_filename_chars(fname, '.xlsx'))
    wb = openpyxl.Workbook()

    # Create the documentation sheets
    ws = wb.active
    ws.title = 'documentation'
    current_row = 1
    for row in xl_doc.split('\n')[2:]:
        cell_addr = f'A{current_row}'
        ws[cell_addr].value = row
        current_row += 1
    # Create other sheets
    for sheet_name, req_data in requireds.items():
        if sheet_name == 'slurry':
            write_slurry_to_excel(wb, pipeline.slurry, req_data)
        elif sheet_name == 'pipeline':
            ws = wb.create_sheet(sheet_name)
            ws[f'A1'] = 'Name:'
            create_and_fill_named_range(wb, sheet_name, 'name', f'B1', pipeline.name)
            write_pipesections_to_excel(wb, pipeline.pipesections, 3)
        else:
            pass
    wb.save(fname)

    wb.close()
    return fname


if __name__ == "__main__":
    import datetime
    from DHLLDV_viewer.load_pump_excel import load_pipeline_from_workbook
    in_fname = "static/pipelines/Example_input.xlsx"
    wb = openpyxl.load_workbook(filename=in_fname, data_only=True)
    PL = load_pipeline_from_workbook(wb)
    PL.name = f'Example Stored Pipeline: {datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")}'
    new_PL_xl_name = store_to_excel(PL, requireds=excel_requireds)

    new_wb = openpyxl.load_workbook(filename=new_PL_xl_name, data_only=True)
    new_PL = load_pipeline_from_workbook(new_wb)
