"""Load pumps and pipelines from excel

The spreadsheet should have sheets with the following properties:
- Names are _not_ case sensitive
- All named ranges have worksheet scope
- All values are in metric units (m, Hz, kW, m3/sec)
- Sheets without 'pipeline', 'pump, 'driver', or 'slurry' in the name are ignored
- One worksheet named 'Pipeline'
    - Has a named range Pipeline!name
    - Has a named range Pipeline!pipe_table with a header row and the following columns:
        - Pipe Name: The name of the pipesection or pump
            - If this is a pipesection, any valid string not containing the word 'pump'
            - If a pump, the name of the tab defining the pump (for pumps the other columns are ignored). Note that if
             you list a pump more than once, the program will make copies, so you can reuse pump definitions.
        - Diameter: The pipe diameter in m
        - Length: The pipe length in m
        - Total K: The total of fitting k-factors for this section of pipe
        - Elev Change: The change in elevation for this section of pipe, in m
- One worksheet named 'Slurry'
    - Has named ranges Slurry!['name', 'Cv', 'd_15', 'd_50', 'd_85', 'fluid', 'pipe_dia', 'rhoi', 'rhos']
    - d_15, d_50, d_85 are diameters in mm
    - pipe_dia is in m
    - fluid is 'salt' or 'fresh'
    - rhos is the solids density in ton/m3
    - rhoi is the insitu density in ton/m3
- One or more worksheets whose name ends in the word 'pump'
    - Has named ranges SomePump!['name', 'design_impeller', 'suction_dia', 'disch_dia','design_speed', 'limited',
    'gear_ratio', 'avail_power','pump_curve']
    - name is any valid python string (be careful of escape sequences)
    - design_impeller, suction_dia, and disch_dia are in m
    - speed is in Hz
    - The limited range can contain the string 'torque', 'power', or 'curve'
    - avail_power is in kW
    - gear_ratio is the engine/motor speed divided by the pump speed
    - The pump_curve range has at least three columns and a header row:
        - Three of the columns have the word 'flow', 'head', or 'power' (not case sensitive) in the first row
        - flow is in m3/sec
        - head is in m.w.c
        - power is in kW
- One or more worksheets matching a pump sheet, with the word pump replaced by the word 'driver' (not case-sensitive)
  in the sheet name. For example, if the pump tab is named "SomePump", then the driver will be 'SomeDriver'.
    - Has named ranges 'SomePump Driver'![name', 'power_curve']
    - name is any valid python string (be careful of escape sequences)
    - power_curve has at least two columns and a header row
        - Two of the columns have the word 'speed' and 'power' in the first row
        - speed is in Hz
        - power is in kW
"""

import copy
import openpyxl
from operator import itemgetter
import warnings

from DHLLDV.DriverObj import Driver
from DHLLDV.PumpObj import Pump
from DHLLDV.PipeObj import Pipe, Pipeline
from DHLLDV.SlurryObj import Slurry
from DHLLDV.DHLLDV_Utils import interpDict

# The following dict defines the required tabs and fields for a valid excel spreadsheet
# The keys are allowed/required tabs
# The values are dicts with one key "required" that is true if the tab is required. The other keys are required defined
# names whose values are the expected type of the contents. If the range is a table, the type is yet another dict, with
# Column names and the type of data in the column
# TODO: The defined names should match the attributes in the relevant objects, to reduce custom parsing
excel_requireds = {'pipeline': {'required': True,
                                'name': str,
                                'pipe_table': {'name': str,
                                               'dia': float,
                                               'length': float,
                                               ('total', 'k'): float,
                                               ('elev', 'change'): float}},
                   'slurry': {'required': True,
                              'name': str,
                              'pipe_dia': float,
                              'd_15': float,
                              'd_50': float,
                              'd_85': float,
                              'fluid': str,
                              'Cv': float,
                              'rhos': float,
                              'rhoi': float
                              },
                   'pump': {'required': False,
                            'name': str,
                            'design_impeller': float,
                            'suction_dia': float,
                            'disch_dia': float,
                            'design_speed': float,
                            'limited': str,
                            'gear_ratio': float,
                            'avail_power': float,
                            'pump_curve': {'flow': float,
                                           'head': float,
                                           'power': float}
                            },
                   'driver': {'required': False,
                              'name': str,
                              'power_curve': {'speed': float,
                                              'power': float}}
                   }


class InvalidExcelError(Exception):
    """The user tried loading an improper Excel worksheet"""


def get_range_value(wb: openpyxl.Workbook, sheet_id: int, range_name: str):
    """Get the value from the given range on the given sheet"""
    sheet_name = wb.sheetnames[sheet_id]
    sheet_addr = wb[sheet_name].defined_names[range_name].attr_text.split("!")[1]
    return wb[sheet_name][sheet_addr].value


def load_driver_from_worksheet(wb: openpyxl.Workbook, sheet_id: int):
    """Create a driver object from a driver worksheet
    wb: The workbook to load
    sheet_id: The id of the driver sheet in the sheet list
    """
    params = {'name': get_range_value(wb, sheet_id, 'name')}
    sheet_name = wb.sheetnames[sheet_id]
    speed_col = None
    power_col = None
    # returns a generator of (worksheet title, cell range) tuples
    cell_range = next(wb[sheet_name].defined_names['power_curve'].destinations)
    rows = wb[sheet_name][cell_range[1]]
    curve = []
    for row in rows:
        vals = [c.value for c in row]
        if speed_col is None:  # Assume the first row is a header
            speed_col = next(i for i, c in enumerate(vals) if 'speed' in c.lower())
            power_col = next(i for i, c in enumerate(vals) if 'power' in c.lower())
        else:
            curve.append((float(vals[speed_col]), float(vals[power_col])))
    curve.sort(key=itemgetter(0))
    params['design_power_curve'] = interpDict(*curve)
    return Driver(**params)


def load_pump_from_worksheet(wb: openpyxl.Workbook, sheet_id: int, driver_id: int = None) -> Pump:
    """Create a pump object from a pump worksheet
    wb: The workbook to load
    sheet_id: The id of the pump sheet in the sheet list
    driver_id: The id of any driver in the sheet list, or None if no driver

    Return the pump object
    """
    single_values = excel_requireds['pump']

    params = dict([(k, v(get_range_value(wb, sheet_id, k)))
                   for k, v in single_values.items() if v in [str, int, float]])
    sheet_name = wb.sheetnames[sheet_id]
    flow_col = None
    head_col = None
    power_col = None
    cell_range = next(wb[sheet_name].defined_names['pump_curve'].destinations)[1]  # returns a generator of (worksheet title, cell range) tuples
    rows = wb[sheet_name][cell_range]
    QH = []
    QP = []
    for row in rows:
        vals = [c.value for c in row]
        if flow_col is None:    # Assume the first row is a header
            flow_col = next(i for i, c in enumerate(vals) if 'flow' in c.lower())
            head_col = next(i for i, c in enumerate(vals) if 'head' in c.lower())
            power_col = next(i for i, c in enumerate(vals) if 'power' in c.lower())
        elif sum([float(vals[flow_col]), float(vals[head_col]), float(vals[power_col])]) == 0:
            pass
        else:
            QH.append((float(vals[flow_col]), float(vals[head_col])))
            QP.append((float(vals[flow_col]), float(vals[power_col])))
    QH.sort(key=itemgetter(0))
    QP.sort(key=itemgetter(0))

    params['design_QH_curve'] = interpDict(*QH)
    params['design_QP_curve'] = interpDict(*QP)
    if QH[-1][0] > 0.0:
        warnings.warn(f'Curves for pump: {params["name"]} do not start at 0, enabling low end interpolation')
        params['design_QH_curve'].extrapolate_low = True
        params['design_QP_curve'].extrapolate_low = True
    if driver_id is not None:
        params['driver'] = load_driver_from_worksheet(wb, driver_id)
        params['driver_name'] = params['driver'].name
    return Pump(**params)


def load_pipeline_from_workbook(wb: openpyxl.Workbook):
    """Create a pipeline object from a workbook
    wb: The workbook to load
    """
    validate_excel(wb)

    pipesheet_id = False
    slurry = False
    pump_sheets = {}
    driver_sheets = {}

    for ws_name in wb.sheetnames:
        sheet_id = wb.sheetnames.index(ws_name)  # The id / index of a worksheet
        if 'pipeline' in ws_name.lower():
            pipeline_name = get_range_value(wb, sheet_id, 'name')
            pipesheet_id = sheet_id
        elif 'pump' in ws_name.lower():
            pump_sheets[ws_name.lower().removesuffix('pump')] = sheet_id
        elif 'driver' in ws_name.lower():
            driver_sheets[ws_name.lower().removesuffix('driver')] = sheet_id
        elif 'slurry' in ws_name.lower():
            slurry = load_slurry_from_workbook(wb, sheet_id)
        else:
            ...

    pumps = {}
    for pump_name in pump_sheets:
        pumps[pump_name] = load_pump_from_worksheet(wb, pump_sheets[pump_name], driver_sheets.get(pump_name))

    sheet_name = wb.sheetnames[pipesheet_id]
    name_col = None
    dia_col = None
    len_col = None
    k_col = None
    dz_col = None
    # returns a generator of (worksheet title, cell range) tuples
    cell_range = next(wb[sheet_name].defined_names['pipe_table'].destinations)[1]
    rows = wb[sheet_name][cell_range]
    pipes = []
    for row in rows:
        vals = [c.value for c in row]
        if name_col is None:  # Assume the first row is a header
            name_col = next(i for i, c in enumerate(vals) if 'name' in c.lower())
            dia_col = next(i for i, c in enumerate(vals) if 'dia' in c.lower())
            len_col = next(i for i, c in enumerate(vals) if 'length' in c.lower())
            k_col = next(i for i, c in enumerate(vals) if all(['total' in c.lower(),
                                                              'k' in c.lower()]))
            dz_col = next(i for i, c in enumerate(vals) if all(['elev' in c.lower(),
                                                                'change' in c.lower()]))
        elif 'pump' in vals[name_col].lower():
            pump_name = vals[name_col].lower().removesuffix('pump')
            pipes.append(copy.copy(pumps[pump_name]))
        else:
            pipes.append(Pipe(name=vals[name_col],
                              diameter=float(vals[dia_col]),
                              length=float(vals[len_col]),
                              total_K=float(vals[k_col]),
                              elev_change=float(vals[dz_col])))
    return Pipeline(name=pipeline_name, pipe_list=pipes, slurry=slurry)


def load_slurry_from_workbook(wb: openpyxl.workbook, sheet_id: int):
    """Load the slurry defined on the 'Slurry' tab of the given workbook

    wb: The workbook with the data
    sheet_id: The id of the Slurry tab"""
    single_values = excel_requireds['slurry']

    params = dict([(k, v(get_range_value(wb, sheet_id, k))) for k, v in single_values.items() if v in [str,
                                                                                                       int,
                                                                                                       float]])
    s = Slurry(name=params['name'],
               Dp=params['pipe_dia'],
               D50=params['d_50']/1000,
               fluid=params['fluid'],
               Cv=params['Cv'],
               )
    s.generate_GSD(d15_ratio=params['d_50']/params['d_15'],
                   d85_ratio=params['d_85']/params['d_50'])
    s.rhos = params['rhos']
    s.rhoi = params['rhoi']
    return s


def validate_excel_fields(wb: openpyxl.workbook, sheet_type: str, sheet_name: int) -> None:
    """Validate that the given sheet has the right fields, raise an InvalidExcelError if not

       wb: The workbook
       sheet_type: The type of sheet (must be a key of excel_requireds)
       sheet_name: The name of the sheet in the wb to check
    """
    sheet_id = wb.sheetnames.index(sheet_name)  # The id / index of a worksheet
    if sheet_type not in excel_requireds.keys():
        raise InvalidExcelError(f'Unknown sheet type {sheet_type = } for {sheet_name = } in {wb.path}, '
                                f'must be a key of load_pump_excel.excel_requireds')
    for field_name, field_type in excel_requireds[sheet_type].items():
        if field_type in [str, float]:
            try:
                value = get_range_value(wb, sheet_id, field_name)
            except AttributeError:
                raise InvalidExcelError(f'Missing field {field_name = } for {sheet_name = } in {wb.path}, '
                                        f'must be worksheet scope')
            if (not type(value) == field_type) and (field_type == float and not type(value) == int):
                raise InvalidExcelError(f'The value of {field_name = } {value = } for {sheet_name = } in {wb.path}, '
                                        f'is of type {type(value)}, should be {field_type}')
        elif isinstance(field_type, dict):
            # returns a generator of (worksheet title, cell range) tuples
            print(f'found {field_name = }')
            try:
                cell_range = next(wb[sheet_name].defined_names[field_name].destinations)[1]
            except KeyError:
                raise InvalidExcelError(f'Did not find {field_name = } {field_type = } for {sheet_name = } in {wb.path}')
            rows = wb[sheet_name][cell_range]
            print(f'{sheet_type = }{sheet_name = } rows: {len(rows)} columns: {len(rows[0])}')
        else:
            print(f'{field_type = } not found for {field_name = }')


def validate_excel(wb: openpyxl.workbook) -> None:
    """Validate that the Excel file has the correct tabs and defined ranges, raise an InvalidExcelError if not
        wb: The workbook
    """
    for sheet_name, fields in excel_requireds.items():
        # First check that the required tabs exist
        present = [s for s in wb.sheetnames if sheet_name in s.lower()]
        if fields['required'] and len(present) != 1:
            raise InvalidExcelError(f'Workbook missing the {sheet_name} tab.')

    for ws_name in wb.sheetnames:
        sheet_type = [t for t in excel_requireds.keys() if t in ws_name.lower()]
        # print(f'validate_excel: Checking {ws_name = } of {sheet_type = }')
        if len(sheet_type) == 1:
            validate_excel_fields(wb, sheet_type[0], ws_name)


if __name__ == "__main__":
    fname = "static/pipelines/Example_input.xlsx"

    wb = openpyxl.load_workbook(filename=fname, data_only=True)  # Loading a workbook, data_only takes the stored values
    validate_excel(wb)
    pumps = {}
    drivers = {}
    for ws_name in wb.sheetnames:
        sheet_id = wb.sheetnames.index(ws_name)       # The id / index of a worksheet
        ret_val = ''
        if 'pipeline' in ws_name.lower():
            input_type = 'pipeline'
            pipeline_name = get_range_value(wb, sheet_id, 'name')
            pipeline = load_pipeline_from_workbook(wb)

    try:
        import sys
        import matplotlib.pyplot as plt
    except ImportError:
        print('matplotlib not found')
        plt = None
        sys.exit()

    fig = plt.figure(figsize=(11, 7.5))
    flow_list = [pipeline.pipesections[-1].flow(v) for v in pipeline.slurry.vls_list]
    flow = 4.0
    for i, pl in enumerate([pipeline]):
        s = pl.slurry
        sp_num = (len(pl.pumps)+1)*100 + 10 + (i + 1)
        head_lists = list(zip(*[pl.calc_system_head(Q) for Q in flow_list]))

        HQ_title = f'{pl.name}: length = {pl.total_length:0.0f} Dp={s.Dp*1000:0.0f}mm, d50={s.D50*1000:0.2f}mm, ' \
                   f'Rsd={s.Rsd:0.3f}, Cv={s.Cv:0.3f}, rhom={s.rhom:0.3f}'
        HQ_plot = fig.add_subplot(sp_num, title=HQ_title, xlim=(0, flow_list[-1]), ylim=(0, head_lists[0][-1]))
        HQ_plot.plot(flow_list, head_lists[0], linewidth=2, linestyle='--', color='black', label="System Cvt=c")
        HQ_plot.plot(flow_list, head_lists[1], linewidth=1, linestyle='--', color='blue', label="System Fluid")
        HQ_plot.plot(flow_list, head_lists[2], linewidth=1, linestyle='-', color='blue', label="Pump Fluid")
        HQ_plot.plot(flow_list, head_lists[3], linewidth=2, linestyle='-', color='black', label="Pump Cvt=c")
        HQ_plot.grid(visible=True, which='both')
        legend = HQ_plot.legend()
        for label in legend.get_texts():
            label.set_fontsize('small')
        for j, pump in enumerate(pl.pumps):
            print(pump.name)
            sp_num += 1
            if pump.limited == 'curve':
                speed_list = [k/pump.gear_ratio for k in pump.driver.design_power_curve.keys()]
            else:
                speed_list = [pump.design_speed*x/10 for x in range(2, 11)]

            req_list = [pump.power_required(flow, n) for n in speed_list]
            avail_list = [pump.power_available(n) for n in speed_list]
            PN_title = f'{pump.driver_name} at Q = {flow:0.3f} m3/sec'
            PN_plot = fig.add_subplot(sp_num, title=PN_title, xlim=(0, speed_list[-1]*1.1))
            PN_plot.plot(speed_list, req_list, linewidth=2, linestyle='--', color='black', label="Power required")
            PN_plot.plot(speed_list, avail_list, linewidth=2, linestyle='-', color='black', label="Pump available")
            PN_plot.grid(visible=True, which='both')
            legend = PN_plot.legend()
            for label in legend.get_texts():
                label.set_fontsize('small')

    plt.tight_layout()
    plt.show()

    print(pipeline.calc_system_head(flow))
    print(pipeline.pumps[1].point(flow))
